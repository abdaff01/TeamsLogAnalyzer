# TeasmLogsV3.py - Version 3.0 - 2024-11-22 - Abdelhay Affoun 

import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
from datetime import datetime
import logging
from typing import Tuple, Dict, Optional
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

class CallStatus:
    SUCCESS = "SUCCESS"
    FAILED = "FAILED"
    WARNING = "WARNING"
    INFO = "INFO"

    @staticmethod
    def get_status_color(status):
        colors = {
            CallStatus.SUCCESS: "90EE90",  # Light green
            CallStatus.FAILED: "FFB6C1",   # Light red
            CallStatus.WARNING: "FFD700",   # Gold
            CallStatus.INFO: "87CEEB"       # Sky blue
        }
        return colors.get(status, "FFFFFF")  # White as default

class SIPCodeAnalyzer:
    def __init__(self):
        self.setup_logging()
        self.load_explanations()

    def setup_logging(self) -> None:
        """Configure logging for the application."""
        log_dir = "logs"
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        log_file = os.path.join(log_dir, f"sip_analyzer_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
        logging.basicConfig(
            filename=log_file,
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )

    def determine_call_status(self, sip_code: int) -> str:
        """Determine call status based on SIP code."""
        if str(sip_code).startswith('2'):
            return CallStatus.SUCCESS
        elif str(sip_code).startswith('1'):
            return CallStatus.INFO
        elif str(sip_code).startswith(('4', '5', '6')):
            return CallStatus.FAILED
        elif str(sip_code).startswith('3'):
            return CallStatus.WARNING
        return CallStatus.INFO

    def format_excel_output(self, writer, df: pd.DataFrame, sheet_name: str) -> None:
        """Format Excel output with colors and styling."""
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Define styles
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format headers
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Adjust column width
            worksheet.column_dimensions[get_column_letter(col_num + 1)].width = max(12, min(30, len(str(value)) + 2))

        # Format data cells and apply conditional formatting
        status_col_index = None
        if 'Status' in df.columns:
            status_col_index = df.columns.get_loc('Status') + 1

        for row_num in range(2, worksheet.max_row + 1):
            # Default color is white if no status column exists
            fill_color = "FFFFFF"
            
            if status_col_index is not None:
                status = worksheet.cell(row=row_num, column=status_col_index).value
                fill_color = CallStatus.get_status_color(status)
            
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Add table style - Replace spaces with underscores in table name
        table_name = f"Table_{sheet_name.replace(' ', '_')}"
        tab = Table(displayName=table_name, ref=worksheet.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)

    def process_file(self, input_path: str, output_path: str) -> Tuple[bool, str]:
        """Process the input Excel file with enhanced formatting and organization."""
        try:
            logging.info(f"Processing file: {input_path}")
            data = pd.read_excel(input_path)

            # Process each call and add status
            results = []
            summaries = []
            
            # Initialize counters for different call types
            total_attempted_calls = 0  # Calls with duration > 0
            successful_calls = 0
            failed_calls = 0
            warning_calls = 0
            info_calls = 0

            for _, row in data.iterrows():
                duration = row.get('Duration (seconds)', 0)  # Updated column name
                
                # Process call status and explanations
                simple_exp, detailed_exp, diagnostic = self.explain_sip(
                    int(row['Final SIP code']),
                    int(row['Final Microsoft subcode']),
                    row['Final SIP Phrase']
                )
                
                status = self.determine_call_status(row['Final SIP code'])
                
                # Count calls based on duration and status
                if duration > 0:
                    total_attempted_calls += 1
                    if status == CallStatus.SUCCESS:
                        successful_calls += 1
                    elif status == CallStatus.FAILED:
                        failed_calls += 1
                    elif status == CallStatus.WARNING:
                        warning_calls += 1
                    elif status == CallStatus.INFO:
                        info_calls += 1

                result = {
                    'Status': status,
                    'Simple_Explanation': simple_exp,
                    'Detailed_Explanation': detailed_exp,
                    'Duration': duration,
                    **diagnostic
                }
                results.append(result)
                
                summary = {
                    'Date': row.get('Start time', 'Unknown'),
                    'User': row.get('Display Name', 'Unknown'),
                    'Status': status,
                    'Duration': duration,
                    'SIP_Code': row['Final SIP code'],
                    'Brief_Explanation': simple_exp
                }
                summaries.append(summary)

            # Create enhanced dataframe
            results_df = pd.DataFrame(results)
            enhanced_data = pd.concat([data, results_df], axis=1)

            # Create summary dataframe
            summary_df = pd.DataFrame(summaries)

            # Generate enhanced statistics
            stats = {
                'Total Attempted Calls (Duration > 0)': total_attempted_calls,
                'Successful Calls': successful_calls,
                'Failed Calls (With Duration)': failed_calls,
                'Warning Calls (With Duration)': warning_calls,
                'Info Calls (With Duration)': info_calls,
                'Average Call Duration (seconds)': round(data['Duration (seconds)'].mean(), 2),  # Updated column name
                'Max Call Duration (seconds)': round(data['Duration (seconds)'].max(), 2),  # Updated column name
                'Success Rate (%)': round((successful_calls / total_attempted_calls * 100 if total_attempted_calls > 0 else 0), 2)
            }
            stats_df = pd.DataFrame(list(stats.items()), columns=['Metric', 'Value'])

            # Save to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Summary sheet
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                self.format_excel_output(writer, summary_df, 'Summary')

                # Detailed Analysis sheet
                enhanced_data.to_excel(writer, sheet_name='Detailed Analysis', index=False)
                self.format_excel_output(writer, enhanced_data, 'Detailed Analysis')

                # Statistics sheet
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)
                self.format_excel_output(writer, stats_df, 'Statistics')

            logging.info(f"Successfully processed file. Output saved to: {output_path}")
            return True, f"Successfully processed file. Output saved to: {output_path}"

        except Exception as e:
            error_msg = f"Error processing file: {str(e)}"
            logging.error(error_msg)
            return False, error_msg

    def load_explanations(self) -> None:
        """Load all SIP code explanations and related data."""
        # First, load the SIP explanations
        self.sip_explanations = {
            # 1xx Provisional Responses
            100: "Call is being processed. Please wait.",
            180: "Phone is ringing at the destination.",
            181: "Call is being redirected to another number.",
            182: "Call is in a queue and will be handled soon.",
            183: "Call setup is in progress.",

            # 2xx Successful Responses
            200: "Call connected successfully.",
            202: "Request accepted and will be processed.",

            # 3xx Redirection Responses
            300: "Multiple call destinations found.",
            301: "Call destination has permanently changed.",
            302: "Call destination is temporarily different.",
            305: "You must use a specific network route.",
            380: "Alternative communication method available.",


            # 4xx Request Failure Responses
            400: "Invalid call request. Check the number.",
            401: "Authentication required to complete the call.",
            403: "Call blocked or not allowed.",
            404: "Phone number or user not found.",
            405: "Calling method not permitted.",
            406: "Call cannot be completed due to incompatible settings.",
            407: "Proxy authentication needed.",
            408: "No response from the destination. Timeout occurred.",
            410: "Number is no longer in service.",
            413: "Call request too large to process.",
            414: "Phone number too complicated to dial.",
            415: "Unsupported communication method.",
            416: "Unrecognized phone number format.",
            420: "Unsupported communication feature.",
            421: "Missing required communication feature.",
            423: "Call setup time too short.",
            480: "Destination temporarily unavailable.",
            481: "Call cannot be found or tracked.",
            482: "Call routing has created a loop.",
            483: "Too many network hops to complete call.",
            484: "Incomplete phone number.",
            485: "Unclear which number to call.",
            486: "Destination is currently busy.",
            487: "Call was cancelled or stopped.",
            488: "Call cannot be accepted by recipient.",

            # 5xx Server Failure Responses
            500: "Network error. Unable to complete call.",
            501: "Call feature not supported.",
            502: "Network routing problem.",
            503: "Network overloaded or maintenance.",
            504: "Network route timeout.",
            505: "Unsupported communication protocol.",
            513: "Call request too large.",

            # 6xx Global Failure Responses
            600: "User is busy everywhere.",
            603: "Call explicitly rejected.",
            604: "User does not exist.",
            606: "Call settings prevent connection."
        }

        # Then, load the Microsoft subcode explanations
        self.microsoft_subcode_explanations = {
            560486: {
                "description": "Network busy or user unavailable",
                "cause": "The destination user agent or network is temporarily unavailable",
                "resolution": "Retry the call after a brief delay. If persistent, check network conditions."
            },
            560487: {
                "description": "Call cancelled or network timeout",
                "cause": "The call was terminated due to timeout or user cancellation",
                "resolution": "Check network latency and connection stability."
            },
            560404: {
                "description": "User or number not found",
                "cause": "The dialed number is invalid or user does not exist",
                "resolution": "Verify the phone number and user existence in the system."
            },
            560480: {
                "description": "Temporary service interruption",
                "cause": "Service is temporarily unavailable",
                "resolution": "Wait for service restoration and retry. Check service status."
            },
            560503: {
                "description": "Service currently unavailable",
                "cause": "System overload or maintenance",
                "resolution": "Wait for service restoration. If persistent, contact support."
            }
        }

    microsoft_subcode_explanations = {
        560486: "Network busy or user unavailable.",
        560487: "Call cancelled or network timeout.",
        560404: "User or number not found.",
        560480: "Temporary service interruption.",
        560503: "Service currently unavailable.",
        0: "No additional details available."
    }

    # Detailed technical explanation
    detailed_explanations = {
        # 1xx Provisional Responses
        100: "Trying - The request is being processed, but no definitive response is available yet.",
        180: "Ringing - The destination user agent is alerting the user.",
        181: "Call Is Being Forwarded - The call is being redirected to another destination.",
        182: "Queued - The request is queued and will be processed soon.",
        183: "Session Progress - Provides progress information about the call setup.",

        # 2xx Successful Responses
        200: "OK - The request has been successfully processed and accepted.",
        202: "Accepted - The request has been accepted for processing, but not completed yet.",

        # 3xx Redirection Responses
        300: "Multiple Choices - The requested address resolves to multiple destinations.",
        301: "Moved Permanently - The requested address is no longer valid and has a new permanent address.",
        302: "Moved Temporarily - The requested address is temporarily unavailable and has a new temporary address.",
        305: "Use Proxy - The client must use the specified proxy to reach the destination.",
        380: "Alternative Service - The request cannot be fulfilled, but an alternative service is available.",

        # 4xx Request Failure Responses
        400: "Bad Request - The request could not be understood due to malformed syntax.",
        401: "Unauthorized - The request requires user authentication.",
        403: "Forbidden - The server understood the request but refuses to authorize it.",
        404: "Not Found - The requested user could not be located on the server.",
        405: "Method Not Allowed - The specified method is not allowed for the requested address.",
        406: "Not Acceptable - The requested resource cannot generate content matching the client's Accept headers.",
        407: "Proxy Authentication Required - The client must first authenticate with the proxy.",
        408: "Request Timeout - No response was received from the destination in a timely manner.",
        410: "Gone - The requested resource is no longer available and will not be available again.",
        413: "Request Entity Too Large - The request payload exceeds server processing capabilities.",
        414: "Request-URI Too Long - The request URI exceeds the server's maximum processing length.",
        415: "Unsupported Media Type - The request includes a media type the server cannot process.",
        416: "Unsupported URI Scheme - The request contains a URI scheme the server does not support.",
        420: "Bad Extension - The server does not understand a specified SIP extension.",
        421: "Extension Required - The server requires a specific extension not present in the request.",
        423: "Interval Too Brief - The request's expiration interval is too short.",
        480: "Temporarily Unavailable - The destination cannot be reached but might be available later.",
        481: "Call/Transaction Does Not Exist - The call or transaction referenced does not exist.",
        482: "Loop Detected - The request indicates a loop in the routing path.",
        483: "Too Many Hops - Maximum number of routing hops has been exceeded.",
        484: "Address Incomplete - The request URI is incomplete.",
        485: "Ambiguous - The request URI is ambiguous and could not be resolved uniquely.",
        486: "Busy Here - The destination is currently busy and cannot accept the call.",
        487: "Request Terminated - The request was terminated by the user or network.",
        488: "Not Acceptable Here - The request cannot be accepted by the recipient.",

        # 5xx Server Failure Responses
        500: "Server Internal Error - An unexpected condition prevented request fulfillment.",
        501: "Not Implemented - The server does not support the functionality required.",
        502: "Bad Gateway - The server received an invalid response from another server.",
        503: "Service Unavailable - The server is temporarily overloaded or under maintenance.",
        504: "Server Time-out - No response received from an upstream server.",
        505: "Version Not Supported - The SIP version is not supported.",
        513: "Message Too Large - The message exceeds the server's processing capabilities.",

        # 6xx Global Failure Responses
        600: "Busy Everywhere - The requested user is busy across all possible locations.",
        603: "Decline - The user explicitly declines the call.",
        604: "Does Not Exist Anywhere - The user cannot be found at any location.",
        606: "Not Acceptable - The user's preferences do not allow the call to be completed."
    }


    # Enhanced Microsoft subcode explanations
    microsoft_subcode_explanations = {
        560486: {
            "description": "Network busy or user unavailable",
            "cause": "The destination user agent or network is temporarily unavailable",
            "resolution": "Retry the call after a brief delay. If persistent, check network conditions."
        },
        560487: {
            "description": "Call cancelled or network timeout",
            "cause": "The call was terminated due to timeout or user cancellation",
            "resolution": "Check network latency and connection stability."
        },
        560404: {
            "description": "User or number not found",
            "cause": "The dialed number is invalid or user does not exist",
            "resolution": "Verify the phone number and user existence in the system."
        },
        560480: {
            "description": "Temporary service interruption",
            "cause": "Service is temporarily unavailable",
            "resolution": "Wait for service restoration and retry. Check service status."
        },
        560503: {
            "description": "Service currently unavailable",
            "cause": "System overload or maintenance",
            "resolution": "Wait for service restoration. If persistent, contact support."
        }
    }

    def explain_sip(self, sip_code: int, microsoft_subcode: int, sip_phrase: str) -> Tuple[str, str, Dict]:
        """
        Explain a SIP code with its Microsoft subcode and phrase.
        Returns: (simple_explanation, detailed_explanation, diagnostic_info)
        """
        # Get simple explanation
        simple_exp = self.sip_explanations.get(sip_code, "Unknown SIP code")
        
        # Get detailed explanation
        detailed_exp = self.detailed_explanations.get(sip_code, "No detailed explanation available")
        
        # Get Microsoft subcode information
        subcode_info = self.microsoft_subcode_explanations.get(microsoft_subcode, {
            "description": "Unknown subcode",
            "cause": "No information available",
            "resolution": "Contact support for more information"
        })
        
        # Create diagnostic dictionary
        diagnostic = {
            "Technical_Details": f"SIP {sip_code} - {sip_phrase}",
            "Microsoft_Subcode": f"{microsoft_subcode} - {subcode_info['description']}",
            "Cause": subcode_info['cause'],
            "Resolution": subcode_info['resolution']
        }
        
        return simple_exp, detailed_exp, diagnostic

class ImprovedGUI:
    def __init__(self):
        self.analyzer = SIPCodeAnalyzer()
        self.create_gui()

    def create_gui(self):
        self.root = tk.Tk()
        self.root.title("Enhanced Call Log Analyzer")
        self.root.geometry("800x600")
        
        # Create main frame with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Add tabs for different sections
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # File processing tab
        file_tab = ttk.Frame(self.notebook, padding="5")
        self.notebook.add(file_tab, text="File Processing")

        # Help/Documentation tab
        help_tab = ttk.Frame(self.notebook, padding="5")
        self.notebook.add(help_tab, text="Help")

        # File selection frame
        file_frame = ttk.LabelFrame(file_tab, text="File Selection", padding="5")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)

        # Input file selection
        self.input_path_var = tk.StringVar()
        ttk.Label(file_frame, text="Input Excel File:").grid(row=0, column=0, padx=5, pady=5)
        ttk.Entry(file_frame, textvariable=self.input_path_var, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_input_file).grid(row=0, column=2, padx=5)

        # Output file selection
        self.output_path_var = tk.StringVar()
        ttk.Label(file_frame, text="Output Excel File:").grid(row=1, column=0, padx=5, pady=5)
        ttk.Entry(file_frame, textvariable=self.output_path_var, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(file_frame, text="Browse", command=self.browse_output_file).grid(row=1, column=2, padx=5)

        # Progress frame
        progress_frame = ttk.LabelFrame(file_tab, text="Progress", padding="5")
        progress_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=5)

        self.progress_var = tk.StringVar(value="Ready to process...")
        ttk.Label(progress_frame, textvariable=self.progress_var).grid(row=0, column=0, padx=5, pady=5)
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress_bar.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=5, pady=5)

        # Process button with icon
        ttk.Button(
            file_tab,
            text="Process Files",
            command=self.process_and_generate,
            style="Accent.TButton"
        ).grid(row=2, column=0, pady=10)

        # Add help content
        help_text = tk.Text(help_tab, wrap=tk.WORD, padx=10, pady=10)
        help_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        help_text.insert(tk.END, """
Call Log Analyzer Help

Color Coding:
• Green: Successful calls (2xx codes)
• Red: Failed calls (4xx, 5xx, 6xx codes)
• Yellow: Warning/Redirected calls (3xx codes)
• Blue: Informational status (1xx codes)

Output Sheets:
1. Summary - Quick overview of all calls
2. Detailed Analysis - Complete call information
3. Statistics - Call statistics and metrics

For more information, please contact support.
        """)
        help_text.config(state=tk.DISABLED)

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=1)
        file_tab.columnconfigure(0, weight=1)
        help_tab.columnconfigure(0, weight=1)
        help_tab.rowconfigure(0, weight=1)

    def browse_input_file(self):
        """Open a file dialog to select an input file."""
        filename = filedialog.askopenfilename(
            title="Select input file",
            filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
        )
        if filename:
            self.input_path_var.set(filename)

    def browse_output_file(self):
        """Open a file dialog to select an output file location."""
        filename = filedialog.asksaveasfilename(
            title="Save output file",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
            defaultextension=".xlsx"
        )
        if filename:
            self.output_path_var.set(filename)

    def process_and_generate(self):
        """Process the input file and generate the output file."""
        input_path = self.input_path_var.get()
        output_path = self.output_path_var.get()
        
        if not input_path or not output_path:
            messagebox.showerror("Error", "Please select both input and output files.")
            return
            
        self.progress_bar.start()
        self.progress_var.set("Processing...")
        
        success, message = self.analyzer.process_file(input_path, output_path)
        
        self.progress_bar.stop()
        if success:
            messagebox.showinfo("Success", message)
        else:
            messagebox.showerror("Error", message)
        self.progress_var.set("Ready to process...")

def main():
    app = ImprovedGUI()
    app.root.mainloop()

if __name__ == "__main__":
    main()